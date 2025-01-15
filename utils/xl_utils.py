import openpyxl


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def get_row_index(file, sheet_name, row_id):
    for i in range(1, get_row_count(file, sheet_name) + 1):
        if read_data(file, sheet_name, i, 1) == row_id:
            return i


def read_data(file, sheet_name, row, column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row, column).value
